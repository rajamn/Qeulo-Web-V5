from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

def format_excel_sheet(ws):
    """
    Apply consistent formatting across all exported Excel sheets:
    - Header styling
    - Freeze top row
    - Auto column widths
    - Alignment
    - Currency formatting where applicable
    """

    # -----------------------
    # Style header row
    # -----------------------
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Apply styles to header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # -----------------------
    # Auto column width
    # -----------------------
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = max(12, length + 2)
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # -----------------------
    # Currency formatting
    # -----------------------
    currency_format = u'₹ #,##0.00'

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                # Apply currency format for probable money columns
                if any(keyword in cell.coordinate for keyword in ["G", "H", "F", "E"]):
                    cell.number_format = currency_format

    # -----------------------
    # Freeze top row
    # -----------------------
    ws.freeze_panes = "A2"

    return ws

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Q
from django.shortcuts import render
from datetime import date, datetime
from appointments.models import AppointmentDetails
from billing.models import PaymentTransaction  # adjust if module name differs
from core.decorators import hospital_admin_required
from openpyxl import Workbook
from .utils import format_excel_sheet
import json
from services.models import Service   # adjust import if needed
from doctors.models import Doctor




@hospital_admin_required
def reports_home(request):
    return render(request, "reports/reports_home.html")

@hospital_admin_required
def daily_opd_report(request):
    hospital = request.user.hospital

    # -------- Date filter (defaults to today) --------
    date_str = request.GET.get("date", "")
    if date_str:
        try:
            report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            report_date = date.today()
    else:
        report_date = date.today()

    # -------- Appointments for that day --------
    appts = (
        AppointmentDetails.objects
        .filter(hospital=hospital, appointment_on=report_date)
        .select_related("doctor", "patient")
    )

    total_reg      = appts.count()
    total_completed = appts.filter(completed=1).count()
    total_queued    = appts.filter(completed=0).count()
    total_cancelled = appts.filter(completed=2).count()

    # -------- Payments for that day --------
    txns = PaymentTransaction.objects.filter(
        hospital=hospital,
        paid_on=report_date,
    )

    total_revenue = txns.exclude(pay_type="Due").aggregate(
        total=Sum("amount")
    )["total"] or 0

    total_due = txns.filter(pay_type="Due").aggregate(
        total=Sum("amount")
    )["total"] or 0

    # Optional: revenue by doctor
    revenue_by_doctor = (
        txns.exclude(pay_type="Due")
        .values("doctor__doctor_name")
        .annotate(total=Sum("amount"))
        .order_by("doctor__doctor_name")
    )

    # -------- Chart Data for Chart.js --------
    chart_data = {
        "status_labels": ["Registered", "In Queue", "Completed", "Cancelled"],
        "status_values": [
            total_reg,
            total_queued,
            total_completed,
            total_cancelled
        ],
        "doctor_labels": [d["doctor__doctor_name"] for d in revenue_by_doctor],
        "doctor_revenue": [float(d["total"]) for d in revenue_by_doctor],
    }


    context = {
        "report_date": report_date,
        "appointments": appts,
        "total_reg": total_reg,
        "total_completed": total_completed,
        "total_queued": total_queued,
        "total_cancelled": total_cancelled,
        "total_revenue": total_revenue,
        "total_due": total_due,
        "revenue_by_doctor": revenue_by_doctor,
        "chart_data_json": json.dumps(chart_data),
    }
    
    return render(request, "reports/daily_opd.html", context)



@hospital_admin_required
def revenue_report(request):
    hospital = request.user.hospital

    # ---------------------------
    # Date range
    # ---------------------------
    start_date = request.GET.get("start")
    end_date = request.GET.get("end")

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today()
    except:
        start = date.today()

    try:
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    except:
        end = date.today()

    qs = PaymentTransaction.objects.filter(
        hospital=hospital,
        paid_on__range=[start, end]
    ).select_related("doctor", "service", "patient")

    # ---------------------------
    # Additional Filters
    # ---------------------------
    doctor_id = request.GET.get("doctor")
    service_id = request.GET.get("service")
    pay_type = request.GET.get("pay_type")

    if doctor_id:
        qs = qs.filter(doctor_id=doctor_id)
    if service_id:
        qs = qs.filter(service_id=service_id)
    if pay_type:
        qs = qs.filter(pay_type=pay_type)

    # ---------------------------
    # KPI Metrics
    # ---------------------------
    total_revenue = qs.exclude(pay_type="Due").aggregate(total=Sum("amount"))["total"] or 0
    total_due = qs.filter(pay_type="Due").aggregate(total=Sum("amount"))["total"] or 0
    total_bills = qs.count()
    avg_revenue = total_revenue / total_bills if total_bills else 0

    # ---------------------------
    # Chart Data
    # ---------------------------
    daily = (
        qs.exclude(pay_type="Due")
        .values("paid_on")
        .annotate(total=Sum("amount"))
        .order_by("paid_on")
    )


    doctor_rev = (
        qs.exclude(pay_type="Due")
          .values("doctor__doctor_name")
          .annotate(total=Sum("amount"))
          .order_by("doctor__doctor_name")
    )

    chart_data = {
        "dates": [d["paid_on"].strftime("%Y-%m-%d") for d in daily],
        "date_values": [float(d["total"]) for d in daily],
        "doctor_labels": [d["doctor__doctor_name"] for d in doctor_rev],
        "doctor_values": [float(d["total"]) for d in doctor_rev],
    }

    context = {
        "transactions": qs.order_by("-paid_on", "-id"),
        "total_revenue": total_revenue,
        "total_due": total_due,
        "total_bills": total_bills,
        "avg_revenue": avg_revenue,
        "start": start,
        "end": end,
        "chart_data_json": json.dumps(chart_data),
        "doctors": Doctor.objects.filter(hospital=hospital),
        "services": Service.objects.filter(hospital=hospital),
    }

    return render(request, "reports/revenue.html", context)


@hospital_admin_required
def revenue_export_excel(request):
    hospital = request.user.hospital

    # ---------------------------
    # Fetch same filters as report
    # ---------------------------
    start_date = request.GET.get("start")
    end_date = request.GET.get("end")

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today()
    except:
        start = date.today()

    try:
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    except:
        end = date.today()

    qs = PaymentTransaction.objects.filter(
        hospital=hospital,
        paid_on__range=[start, end]
    ).select_related("doctor", "service", "patient")

    # Optional filters
    doctor = request.GET.get("doctor")
    service = request.GET.get("service")
    pay_type = request.GET.get("pay_type")

    if doctor:
        qs = qs.filter(doctor_id=doctor)
    if service:
        qs = qs.filter(service_id=service)
    if pay_type:
        qs = qs.filter(pay_type=pay_type)

    # ---------------------------
    # Create Excel Workbook
    # ---------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Report"

    # Header
    headers = ["Date", "Patient", "Doctor", "Service", "Payment Type", "Amount (₹)"]
    ws.append(headers)

    # Rows
    for t in qs:
        ws.append([
            t.paid_on.strftime("%d-%m-%Y"),
            t.patient.patient_name if t.patient else "",
            t.doctor.doctor_name,
            t.service.service_name,
            t.pay_type,
            float(t.amount),
        ])

    # ---------------------------
    # HTTP Response
    # ---------------------------
    format_excel_sheet(ws)

    filename = f"RevenueReport_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@hospital_admin_required
def pending_dues_report(request):
    hospital = request.user.hospital

    # Filters
    start_date = request.GET.get("start")
    end_date = request.GET.get("end")
    doctor_id = request.GET.get("doctor")
    min_due = request.GET.get("min_due", "0")

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today()
    except:
        start = date.today()

    try:
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    except:
        end = date.today()

    # Base queryset — only payments with "Due" transactions
    dues = PaymentTransaction.objects.filter(
        hospital=hospital,
        pay_type="Due",
        paid_on__range=[start, end]
    ).select_related("patient", "doctor", "service", "payment")

    if doctor_id:
        dues = dues.filter(doctor_id=doctor_id)

    if min_due.isdigit():
        dues = dues.filter(amount__gte=int(min_due))

    # KPIs
    total_due = dues.aggregate(total=Sum("amount"))["total"] or 0
    total_items = dues.count()
    avg_due = total_due / total_items if total_items else 0
    max_due = dues.aggregate(max=Sum("amount"))["max"] or 0

    context = {
        "dues": dues.order_by("-paid_on"),
        "total_due": total_due,
        "total_items": total_items,
        "avg_due": avg_due,
        "max_due": max_due,
        "start": start,
        "end": end,
        "doctors": Doctor.objects.filter(hospital=hospital),
    }

    return render(request, "reports/pending_dues.html", context)




@hospital_admin_required
def pending_dues_export_excel(request):
    hospital = request.user.hospital

    start_date = request.GET.get("start")
    end_date = request.GET.get("end")
    doctor_id = request.GET.get("doctor")
    min_due = request.GET.get("min_due", "0")

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today()
    except:
        start = date.today()

    try:
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    except:
        end = date.today()

    dues = PaymentTransaction.objects.filter(
        hospital=hospital,
        pay_type="Due",
        paid_on__range=[start, end]
    ).select_related("patient", "doctor", "service", "payment")

    if doctor_id:
        dues = dues.filter(doctor_id=doctor_id)

    if min_due.isdigit():
        dues = dues.filter(amount__gte=int(min_due))

    # ---------------------------
    # Create Excel workbook
    # ---------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Dues"

    ws.append(["Date", "Patient", "Mobile", "Doctor", "Service", "Due Amount (₹)"])

    for d in dues:
        ws.append([
            d.paid_on.strftime("%d-%m-%Y"),
            d.patient.patient_name if d.patient else "",
            d.patient.contact.mobile_num if d.patient else "",
            d.doctor.doctor_name,
            d.service.service_name,
            float(d.amount),
        ])

    format_excel_sheet(ws)
    filename = f"PendingDues_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



def doctor_productivity_report(request):
    hospital = request.user.hospital

    # Filters
    start_date = request.GET.get("start")
    end_date = request.GET.get("end")

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today()
    except Exception:
        start = date.today()

    try:
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    except Exception:
        end = date.today()

    doctors = Doctor.objects.filter(hospital=hospital)

    report_rows = []
    chart_labels = []
    chart_patients = []
    chart_revenue = []

    for doctor in doctors:
        # Appointments for doctor
        appts = AppointmentDetails.objects.filter(
            hospital=hospital,
            doctor=doctor,
            appointment_on__range=[start, end]
        )

        total_patients = appts.count()
        registered = appts.filter(completed=-1).count()
        queued = appts.filter(completed=0).count()
        completed_count = appts.filter(completed=1).count()
        cancelled = appts.filter(completed=2).count()

        # Revenue for doctor
        revenue_qs = PaymentTransaction.objects.filter(
            hospital=hospital,
            doctor=doctor,
            paid_on__range=[start, end]
        )

        revenue = revenue_qs.exclude(pay_type="Due").aggregate(total=Sum("amount"))["total"] or 0
        due_amount = revenue_qs.filter(pay_type="Due").aggregate(total=Sum("amount"))["total"] or 0

        report_rows.append({
            "doctor": doctor,
            "total_patients": total_patients,
            "registered": registered,
            "queued": queued,
            "completed": completed_count,
            "cancelled": cancelled,
            "revenue": revenue,
            "due_amount": due_amount,
        })

        # Chart data (ignore doctors with 0 activity for neatness)
        if total_patients > 0 or revenue > 0:
            chart_labels.append(doctor.doctor_name)
            chart_patients.append(total_patients)
            chart_revenue.append(float(revenue))

    chart_data = {
        "labels": chart_labels,
        "patients": chart_patients,
        "revenue": chart_revenue,
    }

    context = {
        "rows": report_rows,
        "start": start,
        "end": end,
        "chart_data_json": json.dumps(chart_data),
    }

    return render(request, "reports/doctor_productivity.html", context)


from openpyxl import Workbook
from django.http import HttpResponse


@hospital_admin_required
def doctor_productivity_export_excel(request):
    hospital = request.user.hospital

    # ---------------------------
    # Parse filters
    # ---------------------------
    start_date = request.GET.get("start")
    end_date = request.GET.get("end")

    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today()
    except:
        start = date.today()

    try:
        end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    except:
        end = date.today()

    # ---------------------------
    # Collect data (same as report)
    # ---------------------------
    doctors = Doctor.objects.filter(hospital=hospital)
    export_rows = []

    for doctor in doctors:
        appts = AppointmentDetails.objects.filter(
            hospital=hospital,
            doctor=doctor,
            appointment_on__range=[start, end]
        )

        total_patients = appts.count()
        registered = appts.filter(completed=-1).count()
        queued = appts.filter(completed=0).count()
        completed_count = appts.filter(completed=1).count()
        cancelled = appts.filter(completed=2).count()

        revenue_qs = PaymentTransaction.objects.filter(
            hospital=hospital,
            doctor=doctor,
            paid_on__range=[start, end]
        )

        revenue = revenue_qs.exclude(pay_type="Due").aggregate(total=Sum("amount"))["total"] or 0
        due_amount = revenue_qs.filter(pay_type="Due").aggregate(total=Sum("amount"))["total"] or 0

        export_rows.append([
            doctor.doctor_name,
            total_patients,
            registered,
            queued,
            completed_count,
            cancelled,
            float(revenue),
            float(due_amount),
        ])

    # ---------------------------
    # Create Excel Workbook
    # ---------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Doctor Productivity"

    # Header row
    headers = [
        "Doctor",
        "Total Patients",
        "Registered",
        "In Queue",
        "Completed",
        "Cancelled",
        "Revenue (₹)",
        "Due (₹)"
    ]
    ws.append(headers)

    # Add data rows
    for row in export_rows:
        ws.append(row)
    
    format_excel_sheet(ws)

    # ---------------------------
    # Prepare HTTP Response
    # ---------------------------
    filename = f"DoctorProductivity_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@hospital_admin_required
def waiting_time_report(request):
    hospital = request.user.hospital

    start_date = request.GET.get("start")
    end_date   = request.GET.get("end")

    start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today()
    end   = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()

    qs = AppointmentDetails.objects.filter(
        hospital=hospital,
        appointment_on__range=[start, end],
        completed_at__isnull=False,
        queue_start_time__isnull=False,
    ).select_related("doctor", "patient")

    # ------------------------------------------
    # Compute waiting time (in minutes)
    # ------------------------------------------
    rows = []
    for appt in qs:
        wt_minutes = int((appt.completed_at - appt.queue_start_time).total_seconds() // 60)
        rows.append({
            "doctor": appt.doctor.doctor_name,
            "patient": appt.patient.patient_name,
            "token": appt.token_num,
            "queue_start": appt.queue_start_time,
            "completed_at": appt.completed_at,
            "wait_minutes": wt_minutes,
        })

    # ------------------------------------------
    # Doctor Aggregation
    # ------------------------------------------
    doctor_stats = {}
    for row in rows:
        d = row["doctor"]
        doctor_stats.setdefault(d, {"count": 0, "total_wait": 0})
        doctor_stats[d]["count"] += 1
        doctor_stats[d]["total_wait"] += row["wait_minutes"]

    chart_labels = []
    chart_values = []

    for doctor, stat in doctor_stats.items():
        avg = stat["total_wait"] / stat["count"]
        chart_labels.append(doctor)
        chart_values.append(round(avg, 1))

    chart_data = {
        "labels": chart_labels,
        "values": chart_values,
    }

    return render(request, "reports/waiting_time.html", {
        "rows": rows,
        "start": start,
        "end": end,
        "chart_data_json": json.dumps(chart_data),
    })


@login_required
def reports_home(request):
    hospital = request.user.hospital

    today = date.today()

    # Total OPD today
    opd_today = AppointmentDetails.objects.filter(
        hospital=hospital,
        appointment_on=today
    ).count()

    # Revenue today
    revenue_today = (
        PaymentTransaction.objects.filter(
            hospital=hospital,
            paid_on=today
        )
        .exclude(pay_type="Due")
        .aggregate(total=Sum("amount"))
    )["total"] or 0

    # Pending dues
    pending_dues = (
        PaymentTransaction.objects.filter(
            hospital=hospital,
            pay_type="Due"
        ).aggregate(total=Sum("amount"))
    )["total"] or 0

    # Doctor count
    doctor_count = Doctor.objects.filter(hospital=hospital).count()

    context = {
        "opd_today": opd_today,
        "revenue_today": revenue_today,
        "pending_dues": pending_dues,
        "doctor_count": doctor_count,
    }
    return render(request, "reports/home.html", context)
